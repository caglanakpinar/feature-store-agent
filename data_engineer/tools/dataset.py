"""Readers that turn a data source into one tabular shape, whatever the source happens to be.

A source is either a file — delimited text, JSON, JSON Lines, Parquet, Excel, optionally compressed —
or a table in any database `agent_builder` can connect to (PostgreSQL, MySQL, SQLite, DuckDB,
Snowflake, Redshift, BigQuery). Each of those has its own reader below, with its own arguments and its
own way of getting the bytes off disk or the rows off the wire, and every one of them ends in the same
place: a `Dataset` holding a rectangle of rows keyed by column name.

    from data_engineer.tools.dataset import connector_db, read_dataset, read_delimited, read_table

    data = read_delimited("customers.csv.gz")              # sniffs the delimiter, types the values
    data = read_table("public.customers", db="warehouse")  # a `dbs:` entry in the YAML config
    data = read_table("customers", connection={"db": "postgresql", "host": "...", "database": "crm"})
    data = read_dataset("events.parquet", limit=1000)      # picks the reader by what the source is

    connector = connector_db(db="warehouse")               # connect once, read as often as needed
    data = read_table("customers", connection=connector)

`read_dataset` is the one entry point that routes to the rest, for callers that have a path or a set of
connection details and no reason to care which reader that implies. The individual readers stay public
because they take arguments the router cannot express — a sheet name, a record path, query parameters.
`connector_db` is the connection behind every db read, public for the same reason: a caller reading
several tables should open one connection rather than one per read.

Rows are `list[dict[str, Any]]`, the same shape `BaseSQLDB.query` returns, rather than a dataframe:
nothing here needs pandas, and staying with plain structures keeps a `Dataset` serialisable straight
into a tool response. Optional readers import their driver (`pyarrow`, `openpyxl`) inside the function,
so reading a CSV never requires having installed either.
"""

import bz2
import csv
import gzip
import inspect
import json
import lzma
import re
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from itertools import chain
from pathlib import Path
from typing import Any, Callable, Sequence

from agent_builder import build_sql_db
from uilts.logger import logger


# The config directory a `db` name is looked up in when the caller names one but no `configs`, so
# `read_table(..., db="warehouse")` resolves against this agent's own `dbs:` block.
CONFIG_DIR = Path(__file__).resolve().parent.parent

SUFFIX_FORMATS: dict[str, str] = {  # file extension -> the reader that handles it
    ".csv": "delimited",
    ".tsv": "delimited",
    ".psv": "delimited",
    ".txt": "delimited",
    ".dat": "delimited",
    ".json": "json",
    ".jsonl": "jsonl",
    ".ndjson": "jsonl",
    ".parquet": "parquet",
    ".pq": "parquet",
    ".xlsx": "excel",
    ".xlsm": "excel",
    ".xls": "excel",
}

# Compression is a property of the file, not of its format: `.csv.gz` is read by the delimited reader
# through one of these openers, so every text reader gets compression support for free.
OPENERS: dict[str, Callable[..., Any]] = {".gz": gzip.open, ".bz2": bz2.open, ".xz": lzma.open}

DELIMITERS = ",;\t|"  # what the sniffer is allowed to choose between
SUFFIX_DELIMITERS = {".tsv": "\t", ".psv": "|"}  # a suffix that names its delimiter outright
SAMPLE_BYTES = 64 * 1024  # how much of a file is read to sniff its format and its delimiter

# What an unquoted CSV field is read as. Leading zeros keep a value a string on purpose — "007" and
# "01234" are ids and postcodes far more often than they are numbers, and a coerced one loses its shape.
MISSING_VALUES = frozenset({"", "na", "n/a", "null", "none", "nan", "-", "?"})
BOOLEANS = {"true": True, "false": False}
INTEGER = re.compile(r"^[+-]?(0|[1-9]\d*)$")
DECIMAL = re.compile(r"^[+-]?((0|[1-9]\d*)\.\d*|\.\d+|(0|[1-9]\d*))([eE][+-]?\d+)?$")

# A table or column named in SQL is concatenated into the statement, so it is checked rather than
# escaped: bare identifiers, dotted up to `project.dataset.table` for BigQuery. Anything else is a
# query, and belongs in `read_sql` where the caller writes the SQL themselves.
IDENTIFIER = re.compile(r"^[A-Za-z_][\w$]*(\.[A-Za-z_][\w$]*){0,2}$")


@dataclass
class Dataset:
    """A rectangle of rows, however it was read.

    Every row holds every column — a reader fills in what a source left out — so a consumer can index
    any row by any column name without checking first.

    Args:
        rows: The records, each keyed by column name.
        columns: Column names, in the order the source presented them.
        source: Where the rows came from: a file path, or `engine://database/table` for a db read.
        format: The reader that produced them, e.g. "delimited", "parquet", "sql".
        truncated: True when `limit` cut the read short and the source held more rows.
    """

    rows: list[dict[str, Any]]
    columns: list[str]
    source: str
    format: str
    truncated: bool = False

    def __len__(self) -> int:
        return len(self.rows)

    def head(self, limit: int = 5) -> list[dict[str, Any]]:
        """Return the first `limit` rows, for a look at the data without carrying all of it."""
        return self.rows[:limit]

    def column(self, name: str) -> list[Any]:
        """Return one column's values, in row order."""
        if name not in self.columns:
            raise ValueError(f"{self.source}: no column named {name!r}; it has {self.columns}.")

        return [row.get(name) for row in self.rows]

    def select(self, columns: Sequence[str]) -> "Dataset":
        """Return a copy holding only `columns`, in the order asked for."""
        unknown = [column for column in columns if column not in self.columns]
        if unknown:
            raise ValueError(f"{self.source}: no column(s) {unknown}; it has {self.columns}.")

        return Dataset(
            rows=[{column: row.get(column) for column in columns} for row in self.rows],
            columns=list(columns),
            source=self.source,
            format=self.format,
            truncated=self.truncated,
        )

    def to_dict(self, limit: int | None = None) -> dict[str, Any]:
        """Render as JSON-safe structures — what a tool hands back to an agent.

        Values a driver returns that JSON has no type for (dates, `Decimal`, `bytes`) are converted
        here rather than at read time, so the rows themselves stay in the types they were read as.
        """
        rows = self.rows if limit is None else self.rows[:limit]
        return {
            "source": self.source,
            "format": self.format,
            "row_count": len(self.rows),
            "column_count": len(self.columns),
            "columns": self.columns,
            "truncated": self.truncated,
            "rows": [{key: _json_safe(value) for key, value in row.items()} for row in rows],
        }


# ---- files ------------------------------------------------------------------------------------------


def read_delimited(
    path: str | Path,
    delimiter: str | None = None,
    encoding: str = "utf-8",
    has_header: bool = True,
    column_names: Sequence[str] | None = None,
    limit: int | None = None,
    columns: Sequence[str] | None = None,
    coerce: bool = True,
    missing_values: Sequence[str] | None = None,
) -> Dataset:
    """Read delimited text — CSV, TSV, pipe-separated, and the same compressed.

    The delimiter is taken from `delimiter`, else from a suffix that names one (`.tsv`, `.psv`), else
    sniffed from the head of the file. Rows are read streaming and stop at `limit`, so a preview of a
    large file costs the preview rather than the file.

    Args:
        path: File to read; `.gz`, `.bz2` and `.xz` are decompressed on the way in.
        delimiter: Field separator. Sniffed when not given.
        encoding: Text encoding of the file.
        has_header: Whether the first row names the columns.
        column_names: Column names to use, required when `has_header` is False unless positional
            `column_1..column_n` names are acceptable.
        limit: Stop after this many rows.
        columns: Keep only these columns.
        coerce: Read unquoted fields as numbers, booleans and nulls instead of leaving every value a
            string — a delimited file carries no types of its own.
        missing_values: Strings to read as None. Defaults to `MISSING_VALUES`.

    Returns:
        The rows, with the columns the header (or `column_names`) declared.
    """
    path = Path(path)
    missing = frozenset(value.lower() for value in (missing_values or MISSING_VALUES))
    delimiter = delimiter or _sniff_delimiter(path, encoding)

    with _open_text(path, encoding) as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        names = list(column_names) if column_names else []
        if has_header:
            header = next(reader, [])
            names = names or _unique_names(header)

        records: list[dict[str, Any]] = []
        ragged = 0
        truncated = False
        for values in reader:
            if not values or values == ['']:  # a blank line between records, not a row of empties
                continue
            if not names:  # no header and no names given: the first row's width decides them
                names = _unique_names([''] * len(values))
            if limit is not None and len(records) >= limit:
                truncated = True
                break

            if len(values) != len(names):
                ragged += 1
            fields: list[Any] = list(values[: len(names)])
            fields += [None] * (len(names) - len(fields))
            if coerce:
                fields = [_coerce(value, missing) for value in fields]
            records.append(dict(zip(names, fields)))

    if ragged:
        logger.warning(f"{path}: {ragged} row(s) did not have {len(names)} fields; padded or trimmed.")

    return _dataset(records, str(path), "delimited", names, columns, truncated)


def read_json(
    path: str | Path,
    encoding: str = "utf-8",
    record_path: str | None = None,
    limit: int | None = None,
    columns: Sequence[str] | None = None,
) -> Dataset:
    """Read a JSON document holding records.

    An array of objects is the records. An object is looked into: `record_path` names the key holding
    them (dotted, e.g. `"result.items"`); with no `record_path`, a lone array value is taken as the
    records, and an object with none is read as a single row.

    Args:
        path: File to read; compressed extensions are handled as in `read_delimited`.
        encoding: Text encoding of the file.
        record_path: Dotted path to the array of records inside the document.
        limit: Keep only the first this many records.
        columns: Keep only these columns.

    Returns:
        The rows, with the union of the keys the records carried as columns.
    """
    path = Path(path)
    with _open_text(path, encoding) as handle:
        document = json.load(handle)

    records = _records_in(document, record_path, str(path))
    truncated = limit is not None and len(records) > limit
    if limit is not None:
        records = records[:limit]

    return _dataset(records, str(path), "json", None, columns, truncated)


def read_jsonl(
    path: str | Path,
    encoding: str = "utf-8",
    limit: int | None = None,
    columns: Sequence[str] | None = None,
    skip_invalid: bool = False,
) -> Dataset:
    """Read JSON Lines — one JSON object per line — a line at a time, stopping at `limit`.

    Args:
        path: File to read; compressed extensions are handled as in `read_delimited`.
        encoding: Text encoding of the file.
        limit: Stop after this many records.
        columns: Keep only these columns.
        skip_invalid: Log and skip a line that does not parse, instead of raising. Export dumps are
            often truncated mid-line at the end, which this tolerates.

    Returns:
        The rows, with the union of the keys the records carried as columns.
    """
    path = Path(path)
    records: list[dict[str, Any]] = []
    truncated = False

    with _open_text(path, encoding) as handle:
        for number, line in enumerate(handle, start=1):
            line = line.strip()
            if not line:
                continue
            if limit is not None and len(records) >= limit:
                truncated = True
                break

            try:
                record = json.loads(line)
            except json.JSONDecodeError as error:
                if not skip_invalid:
                    raise ValueError(f"{path}: line {number} is not valid JSON ({error}).") from error
                logger.warning(f"{path}: skipping line {number}, which is not valid JSON ({error}).")
                continue

            if not isinstance(record, dict):
                record = {"value": record}
            records.append(record)

    return _dataset(records, str(path), "jsonl", None, columns, truncated)


def read_parquet(
    path: str | Path,
    limit: int | None = None,
    columns: Sequence[str] | None = None,
) -> Dataset:
    """Read a Parquet file, via `pyarrow`.

    Column selection is pushed down to the file — Parquet is columnar, so unread columns are never
    touched — and a `limit` reads batches rather than the whole file.

    Args:
        path: File to read.
        limit: Stop after this many rows.
        columns: Keep only these columns; read without touching the others.

    Returns:
        The rows, in the types Parquet stored them as.
    """
    import pyarrow.parquet as parquet  # a heavy dependency only this reader needs

    path = Path(path)
    file = parquet.ParquetFile(path)
    try:
        names = list(file.schema_arrow.names)
        wanted = list(columns) if columns else None
        if wanted:
            unknown = [column for column in wanted if column not in names]
            if unknown:
                raise ValueError(f"{path}: no column(s) {unknown}; it has {names}.")

        if limit is None:
            records = file.read(columns=wanted).to_pylist()
            truncated = False
        else:
            records = []
            for batch in file.iter_batches(batch_size=max(1, min(limit, 65536)), columns=wanted):
                records.extend(batch.to_pylist())
                if len(records) >= limit:
                    break
            truncated = file.metadata.num_rows > limit
            records = records[:limit]
    finally:
        file.close()

    return _dataset(records, str(path), "parquet", wanted or names, None, truncated)


def read_excel(
    path: str | Path,
    sheet: str | int | None = None,
    has_header: bool = True,
    limit: int | None = None,
    columns: Sequence[str] | None = None,
    coerce: bool = False,
) -> Dataset:
    """Read one worksheet of an Excel workbook, via `openpyxl`.

    Opened read-only with formulas resolved to their last computed value, so a workbook is streamed
    rather than loaded whole and a formula cell reads as the number it evaluated to.

    Args:
        path: Workbook to read.
        sheet: Sheet name, or its index in the workbook. Defaults to the active sheet.
        has_header: Whether the first row names the columns.
        limit: Stop after this many rows.
        columns: Keep only these columns.
        coerce: Read text cells as numbers, booleans and nulls. Off by default — Excel cells already
            carry types, so text here is usually text on purpose.

    Returns:
        The rows, in the types the workbook stored them as.
    """
    from openpyxl import load_workbook  # only this reader needs it

    path = Path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet is None:
            worksheet = workbook.active
        elif isinstance(sheet, int):
            worksheet = workbook.worksheets[sheet]
        elif sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
        else:
            raise ValueError(f"{path}: no sheet named {sheet!r}; it has {workbook.sheetnames}.")

        rows = worksheet.iter_rows(values_only=True)
        first = next(rows, ())
        if has_header:
            names = _unique_names([str(cell) if cell is not None else '' for cell in first])
        else:  # the row we looked at is data: it sets the width, and goes back on the front
            names = _unique_names([''] * len(first))
            rows = chain([first], rows)

        records: list[dict[str, Any]] = []
        truncated = False
        for values in rows:
            if all(value is None for value in values):  # trailing blank rows the sheet's extent keeps
                continue
            if limit is not None and len(records) >= limit:
                truncated = True
                break

            values = list(values[: len(names)]) + [None] * (len(names) - len(values))
            if coerce:
                values = [_coerce(value, MISSING_VALUES) for value in values]
            records.append(dict(zip(names, values)))
    finally:
        workbook.close()

    return _dataset(records, f"{path}::{worksheet.title}", "excel", names, columns, truncated)


# ---- databases --------------------------------------------------------------------------------------


def connector_db(
    db: str | None = None,
    connection: dict[str, Any] | Any = None,
    configs: Any = None,
    settings: dict[str, Any] | None = None,
) -> Any:
    """Open a connection to any supported database, from a config name or from connection details.

    This is what the readers below connect with, public because a caller reading more than once should
    connect once: hand what this returns back as `connection=` and every read reuses it, instead of
    each one opening and closing a connection of its own. Closing it is then the caller's to do.

        connector = connector_db(db="warehouse")  # a `dbs:` entry in the YAML config
        try:
            customers = read_table("customers", connection=connector)
            orders = read_sql("SELECT * FROM orders WHERE created_at > ?", connection=connector,
                              params=("2026-01-01",))
        finally:
            connector.close()

    Args:
        db: Name of a `dbs:` entry to connect with. Its `password` and `credentials` are read as env
            var names when they name one, so a secret need not be passed through here at all.
        connection: Connection details for a database no config describes — `db` names the engine
            ("postgresql", "mysql", "sqlite", "duckdb", "snowflake", "redshift", "bigquery") and the
            rest are that engine's fields (`host`, `port`, `database`, `user`, `password`,
            `connection_string`, `path`, `project`, `credentials`). An already-open connector passed
            here is returned as it is, so this is safe to call on whatever a caller happens to hold.
        configs: Config directory or `Configs` to resolve `db` in. Defaults to this agent's config.
        settings: Driver-specific options the engine takes, e.g. `{"sslmode": "require"}` for
            PostgreSQL, `{"charset": "utf8mb4"}` for MySQL, `{"warehouse": "..."}` for Snowflake.

    Returns:
        A connected `BaseSQLDB`: `query(sql, params)` returns rows as dicts, `execute(sql, params)`
        returns the affected row count, and `close()` ends the connection.
    """
    connector, _ = _connect(db, connection, configs, settings)
    return connector


def read_sql(
    query: str,
    db: str | None = None,
    connection: dict[str, Any] | Any = None,
    configs: Any = None,
    params: Sequence[Any] | dict[str, Any] | None = None,
    settings: dict[str, Any] | None = None,
    limit: int | None = None,
    columns: Sequence[str] | None = None,
) -> Dataset:
    """Run a query against any supported database and return its rows.

    The connection comes from one of three places: `db`, naming an entry in the config's `dbs:` block;
    `connection`, a mapping of connection details (`db` names the engine — "postgresql", "bigquery",
    "sqlite", …) for a database no config describes; or `connection` as an already-open connector from
    `connector_db`, which is then reused and left open for the caller to close.

    `limit` truncates the result after it is fetched rather than editing the SQL, since the statement
    belongs to the caller — pass a LIMIT in the query itself to keep the rows off the wire.

    Args:
        query: SQL to run.
        db: Name of a `dbs:` entry to connect with.
        connection: Connection details, or an already-open `BaseSQLDB`.
        configs: Config directory or `Configs` to resolve `db` in. Defaults to this agent's config.
        params: Bound parameters for the query, in the driver's placeholder style.
        settings: Driver-specific options, e.g. `{"sslmode": "require"}`.
        limit: Keep only the first this many rows.
        columns: Keep only these columns.

    Returns:
        The rows the query returned, in the types the driver produced.
    """
    connector, opened = _connect(db, connection, configs, settings)
    try:
        records = connector.query(query, params)
    finally:
        if opened:  # a connector the caller handed in stays theirs to close
            connector.close()

    truncated = limit is not None and len(records) > limit
    if limit is not None:
        records = records[:limit]

    return _dataset(records, _source_of(connector), "sql", None, columns, truncated)


def read_table(
    table: str,
    db: str | None = None,
    connection: dict[str, Any] | Any = None,
    configs: Any = None,
    settings: dict[str, Any] | None = None,
    limit: int | None = None,
    columns: Sequence[str] | None = None,
) -> Dataset:
    """Read a whole table, or the first `limit` rows of it, without writing the SQL.

    Connection arguments are `read_sql`'s. The statement is built here, so `table` and `columns` are
    validated as identifiers rather than escaped — anything more involved than a table and a row cap
    is a query, and `read_sql` takes those.

    Args:
        table: Table to read, optionally qualified: `schema.table`, or `project.dataset.table`.
        db: Name of a `dbs:` entry to connect with.
        connection: Connection details, or an already-open `BaseSQLDB`.
        configs: Config directory or `Configs` to resolve `db` in. Defaults to this agent's config.
        settings: Driver-specific options, e.g. `{"sslmode": "require"}`.
        limit: Read at most this many rows — a LIMIT in the statement, so the rest never leaves the db.
        columns: Read only these columns.

    Returns:
        The table's rows, in the types the driver produced.
    """
    for identifier in [table, *(columns or [])]:
        if not IDENTIFIER.match(str(identifier)):
            raise ValueError(
                f"{identifier!r} is not a plain identifier; pass the statement to `read_sql` instead."
            )

    selection = ", ".join(columns) if columns else "*"
    query = f"SELECT {selection} FROM {table}"
    if limit is not None:
        query += f" LIMIT {int(limit) + 1}"  # one over, so a full page can be told from a full table

    data = read_sql(query, db=db, connection=connection, configs=configs, settings=settings)
    truncated = limit is not None and len(data.rows) > limit
    return Dataset(
        rows=data.rows[:limit] if limit is not None else data.rows,
        columns=data.columns,
        source=f"{data.source}/{table}",
        format="sql",
        truncated=truncated,
    )


# ---- routing ----------------------------------------------------------------------------------------


READERS: dict[str, Callable[..., Dataset]] = {
    "delimited": read_delimited,
    "csv": read_delimited,
    "json": read_json,
    "jsonl": read_jsonl,
    "parquet": read_parquet,
    "excel": read_excel,
}


def read_dataset(
    data_source: str | Path | None = None,
    format: str | None = None,
    query: str | None = None,
    table: str | None = None,
    db: str | None = None,
    connection: dict[str, Any] | Any = None,
    configs: Any = None,
    limit: int | None = None,
    columns: Sequence[str] | None = None,
    **options: Any,
) -> Dataset:
    """Read whatever the caller has — a path or a set of connection details — into a `Dataset`.

    Routing, in order: a `query` runs; a `table` is read; anything else is a file, whose reader comes
    from `format`, else from its extension, else from a look at its first bytes. `data_source` doubles
    as the table name or the query when one is passed there rather than in its own argument, which is
    what lets a single tool argument carry either kind of source.

    Args:
        data_source: File path, table name, or SQL — whichever this source is.
        format: Reader to use, overriding what the file looks like: "delimited", "json", "jsonl",
            "parquet", "excel", or "sql".
        query: SQL to run instead of reading a file.
        table: Table to read instead of reading a file.
        db: Name of a `dbs:` entry to connect with.
        connection: Connection details, or an already-open `BaseSQLDB`.
        configs: Config directory or `Configs` to resolve `db` in. Defaults to this agent's config.
        limit: Stop after this many rows.
        columns: Keep only these columns.
        **options: Passed on to the chosen reader — `delimiter`, `sheet`, `record_path`, `encoding`,
            `has_header`, `coerce`, … Anything that reader does not take is dropped with a warning,
            so one caller can pass the options for a format it does not know it has.

    Returns:
        The rows the source held, however it had to be read.
    """
    if query or (format == "sql" and data_source):
        return read_sql(
            query or str(data_source), db=db, connection=connection, configs=configs,
            limit=limit, columns=columns, **_accepted(read_sql, options),
        )

    if table or db or connection:
        name = table or data_source
        if not name:
            raise ValueError("a database read needs a `table` to read or a `query` to run.")
        return read_table(
            str(name), db=db, connection=connection, configs=configs,
            limit=limit, columns=columns, **_accepted(read_table, options),
        )

    if not data_source:
        raise ValueError("nothing to read: pass a `data_source`, or a `table`/`query` and a db.")

    path = Path(data_source)
    if not path.exists():
        raise FileNotFoundError(f"no such data source: {path}")

    reader_name = format or detect_format(path)
    reader = READERS.get(reader_name)
    if not reader:
        raise ValueError(f"unknown format {reader_name!r}; expected one of {sorted(READERS)}.")

    logger.info(f"Reading {path} as {reader_name}.")
    return reader(path, limit=limit, columns=columns, **_accepted(reader, options))


def detect_format(path: str | Path) -> str:
    """Work out which reader a file needs, from its extension or else from its first bytes.

    The extension decides it when there is one. Otherwise the head of the file does: Parquet's magic
    number, a line that parses as a JSON object (one per line being JSON Lines, one document being
    JSON), and delimited text as the fallback, which is what an extensionless export usually is.
    """
    path = Path(path)
    suffixes = [suffix.lower() for suffix in path.suffixes]
    for suffix in reversed(suffixes):  # `.csv.gz` is delimited: the format is under the compression
        if suffix in OPENERS:
            continue
        if suffix in SUFFIX_FORMATS:
            return SUFFIX_FORMATS[suffix]
        break

    with _open_binary(path) as handle:
        head = handle.read(SAMPLE_BYTES)

    if head[:4] == b"PAR1":
        return "parquet"
    if head[:2] == b"PK":  # a zip container, which is what an .xlsx workbook is
        return "excel"

    lines = [line.strip() for line in head.decode("utf-8", errors="replace").splitlines() if line.strip()]
    if lines and lines[0][:1] in "{[":
        if lines[0].startswith("{") and _parses_as_object(lines[0]):
            return "jsonl"  # a whole document would not have closed on its first line
        return "json"

    return "delimited"


# ---- internals --------------------------------------------------------------------------------------


def _open_text(path: Path, encoding: str) -> Any:
    """Open a text file, decompressing by extension. `newline=''` is what the csv module requires."""
    opener = OPENERS.get(path.suffix.lower(), open)
    return opener(path, mode="rt", encoding=encoding, newline='')


def _open_binary(path: Path) -> Any:
    """Open a file's bytes, decompressing by extension."""
    opener = OPENERS.get(path.suffix.lower(), open)
    return opener(path, mode="rb")


def _sniff_delimiter(path: Path, encoding: str) -> str:
    """Pick the field separator: named by the extension, else sniffed, else a comma."""
    for suffix in reversed([suffix.lower() for suffix in path.suffixes]):
        if suffix in OPENERS:
            continue
        if suffix in SUFFIX_DELIMITERS:
            return SUFFIX_DELIMITERS[suffix]
        break

    with _open_text(path, encoding) as handle:
        sample = handle.read(SAMPLE_BYTES)

    try:
        return csv.Sniffer().sniff(sample, delimiters=DELIMITERS).delimiter
    except csv.Error:  # too uniform to sniff — one column, or a single row
        header = sample.splitlines()[0] if sample else ''
        counts = {delimiter: header.count(delimiter) for delimiter in DELIMITERS}
        best = max(counts, key=lambda delimiter: counts[delimiter])
        logger.info(f"{path}: could not sniff a delimiter; using {best!r}.")
        return best if counts[best] else ","


def _unique_names(header: Sequence[Any]) -> list[str]:
    """Turn a header row into usable column names: blanks filled, duplicates suffixed."""
    seen: dict[str, int] = {}
    names: list[str] = []
    for position, name in enumerate(header, start=1):
        base = (str(name).strip() if name is not None else '') or f"column_{position}"
        occurrence = seen.get(base, 0) + 1
        name = base if occurrence == 1 else f"{base}_{occurrence}"
        while name in seen:  # the suffixed name may already be a column in its own right
            occurrence += 1
            name = f"{base}_{occurrence}"

        seen[base] = occurrence
        seen[name] = occurrence
        names.append(name)

    return names


def _coerce(value: Any, missing: frozenset[str]) -> Any:
    """Read one text field as the value it represents; anything else is passed through untouched."""
    if not isinstance(value, str):
        return value

    text = value.strip()
    if text.lower() in missing:
        return None
    if text.lower() in BOOLEANS:
        return BOOLEANS[text.lower()]
    if INTEGER.match(text):
        return int(text)
    if DECIMAL.match(text):
        return float(text)

    return text


def _records_in(document: Any, record_path: str | None, source: str) -> list[dict[str, Any]]:
    """Find the records inside a loaded JSON document. See `read_json` for the rules."""
    if record_path:
        for key in record_path.split("."):
            if not isinstance(document, dict) or key not in document:
                raise ValueError(f"{source}: no {record_path!r} in the document.")
            document = document[key]

    if isinstance(document, dict):
        arrays = [value for value in document.values() if isinstance(value, list)]
        document = arrays[0] if len(arrays) == 1 else [document]

    if not isinstance(document, list):
        raise ValueError(f"{source}: holds a {type(document).__name__}, not records.")

    return [record if isinstance(record, dict) else {"value": record} for record in document]


def _dataset(
    records: list[dict[str, Any]],
    source: str,
    format: str,
    names: Sequence[str] | None,
    columns: Sequence[str] | None,
    truncated: bool,
) -> Dataset:
    """Square the records off into a rectangle — the one shape every reader above ends in.

    `names` is the column order a source declared (a header, a schema); without one the order is the
    order keys were first seen. Either way every row is filled out to hold every column, so a caller
    never has to ask whether a row has a field.
    """
    if names is None:
        names = list(dict.fromkeys(key for record in records for key in record))

    rows = [{name: record.get(name) for name in names} for record in records]
    data = Dataset(rows=rows, columns=list(names), source=source, format=format, truncated=truncated)
    logger.info(f"Read {len(rows)} row(s) and {len(data.columns)} column(s) from {source}.")

    return data.select(columns) if columns else data


def _connect(
    db: str | None,
    connection: dict[str, Any] | Any,
    configs: Any,
    settings: dict[str, Any] | None,
) -> tuple[Any, bool]:
    """Resolve the three ways a database is named into `(connector, we_opened_it)`.

    The ownership half is why this sits behind `connector_db`: a connector opened here is closed by
    the reader that opened it, and one the caller handed in is left alone.
    """
    if connection is not None and hasattr(connection, "query"):  # already open; reuse as it is
        return connection, False

    if not db and not connection:
        raise ValueError("no database to read from: pass `db` or `connection`.")

    return build_sql_db(
        name=db,
        configs=configs or (CONFIG_DIR if db else None),
        settings=settings,
        **(connection or {}),
    ), True


def _source_of(connector: Any) -> str:
    """Name where a db read came from, as `engine://database`.

    Read with `getattr` rather than off the fields directly, because the connector may be one the
    caller opened themselves and handed in — anything with a `query` method is allowed to be one.
    """
    engine = getattr(connector, "db", None) or "sql"
    target = (
        getattr(connector, "database", None)
        or getattr(connector, "path", None)
        or getattr(connector, "name", None)
        or "db"
    )
    return f"{engine}://{target}"


def _accepted(reader: Callable[..., Any], options: dict[str, Any]) -> dict[str, Any]:
    """Keep only the options `reader` takes, so one router can carry every reader's arguments."""
    parameters = inspect.signature(reader).parameters
    accepted = {name: value for name, value in options.items() if name in parameters}
    for name in options.keys() - accepted.keys():
        logger.warning(f"{reader.__name__} does not take {name!r}; dropping it.")

    return accepted


def _parses_as_object(line: str) -> bool:
    """Whether one line is a complete JSON object — how JSON Lines is told from a JSON document."""
    try:
        return isinstance(json.loads(line), dict)
    except json.JSONDecodeError:
        return False


def _json_safe(value: Any) -> Any:
    """Convert a driver's value into something `json.dumps` accepts."""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (bytes, bytearray, memoryview)):
        return bytes(value).hex()
    if isinstance(value, (list, tuple, set)):
        return [_json_safe(item) for item in value]
    if isinstance(value, dict):
        return {str(key): _json_safe(item) for key, item in value.items()}

    return str(value)
